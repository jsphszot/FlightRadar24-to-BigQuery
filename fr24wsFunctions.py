# × pac-miami@ed-cm-caranalytics-dev.iam.gserviceaccount.com
# ed-cm-caranalytics-dev@appspot.gserviceaccount.com

# region: Time Functions and Variables --------------------------------------------------

import datetime as dt
import pytz
class processTime():
    """
    startTime: a datetime object (usually datatime.now()).
    format: default 's', returns total seconds, 'f' returns min and secs.
    startTime, format='s'
    """
    def __init__(self):
        self.startTimer = dt.datetime.now()
        self.miamitime = dt.datetime.now(pytz.timezone('US/Eastern')).strftime("%Y-%m-%d %H:%M:%S") #dt.datetime.utcnow()
        self.NowTimeStamp = int(dt.datetime.timestamp(self.startTimer))
        self.todayfrmtd = self.startTimer.strftime("%Y-%m-%d")
        self.scrapedateid = self.startTimer.strftime("%Y-%m-%d %H:%M:%S")
        self.monthDay = self.startTimer.strftime('%Y-%b-%d')
        self.hourMins = self.startTimer.strftime('%H:%M')
        self.intDay = int(self.startTimer.strftime('%d'))
        self.thisWeek = int(self.startTimer.isocalendar()[1])

    def TotalRunningTime(self, format='s'):
        """
        Spits out total running time, default is 's' in seconds.\n
        Any other value in format will return mins and seconds.
        """
        totalTime = round((dt.datetime.now() - self.startTimer).total_seconds(), 2)
        ftotalTime = "{}m and {}s".format(int(totalTime//60), int(totalTime % 60))
        if format == 's':
            return totalTime
        else:
            return ftotalTime

    def SubProcRunTime(self, SubProcStartTime, format='s'):
        subptotalTime=round((dt.datetime.now() - SubProcStartTime).total_seconds(), 2)
        ftotalTime = "{}m and {}s".format(int(subptotalTime//60), int(subptotalTime % 60))
        if format == 's':
            return subptotalTime
        else:
            return ftotalTime


class DictTime:
    def __init__(self):
        self.dictTime = {1:'LUN',2:'MAR',3:'MIE',4:'JUE',5:'VIE',6:'SAB',7:'DOM',}

    def DoW(self, Date):
        """
        returns Name of Day of week (LUN-DOM)
        Date input as timestamp
        """
        isoWeekDay = Date.isoweekday()
        return self.dictTime.get(isoWeekDay)

    def Tstmp2Str(self, Date, Format="ymd"):
        """
        parses timestamp to given format, can be ymd, ymdhms, or hms
        Date input as timestamp
        """
        inputedFormat={"ymd": "%Y-%m-%d", 'ymdhms': "%Y-%m-%d %H:%M:%S", "hms": "%H:%M:%S"}.get(Format)

        try:
            return Date.strftime(inputedFormat)
        except:
            return None

    def WeekNum(self, Date):
        """
        returns isocalendar week number
        Date input as timestamp
        """
        return int(Date.isocalendar()[1]) # weeknum of dep lt

    def weeks_for_year(self, year):
        """
        Returns number of weeks for a given year
        """
        last_week = dt.datetime(year, 12, 28)
        return last_week.isocalendar()[1]

    def RelativeWeek(self, date):
        """
        Defines weeks from LY as negative distance from current year.\n
        For example:\n
        \tlast week 2019 = 52. Relative to 2020 it would be week 0.
        \tweek 51 year 2019 = week -1 2020
        """
        results_to_return = []
        dttoday=dt.datetime.today()
        thisyear=dttoday.isocalendar()[0]
        for ldate in date:
            lldate=dt.datetime.strptime(ldate, "%Y-%m-%d %H:%M:%S")
            llyear, llweek, *_ = lldate.isocalendar()
            # subtract_weeks=(thisyear-llyear)*weeks_for_year(llyear)
            subtract_weeks=sum([weeks_for_year(x) for x in list(range(llyear, thisyear))])
            result_to_append=llweek-subtract_weeks
            results_to_return.append(result_to_append)
        return results_to_return


import re
def reTime(x):
    """
    wrapper for re.search that finds time HH:MM element in string
    """
    foundem=re.search('\d{2}:\d{2}', x)
    return foundem.group(0) if foundem else "-"

def reStatus(x):
    """
    wrapper for re.search that finds [A-Z|a-z]+ element in string
    """
    foundem=re.search('[A-Z|a-z]+', x)
    return foundem.group(0) if foundem else "-"   

def ToTimeIfPossible(Timestamp):
    """
    Formats datetime timestamp as '%Y-%m-%d %H:%M:%S' if possible, otherwise applies None type.
    """
    return None if Timestamp == None else dt.datetime.fromtimestamp(Timestamp).strftime('%Y-%m-%d %H:%M:%S') 

# endregion
# ---------------------------------------------------------------------------------------

# region: Get info from GoogleSheets ----------------------------------------------------
# https://docs.google.com/spreadsheets/d/1y-3K9e47GWZTt3iPAH4tgRfey2IgZCp2a1-FD5adSQ8/edit?usp=sharing 
# https://gspread.readthedocs.io/en/latest/oauth2.html

class GetGoogleSheetInfo:
    """
    WhichProcess arg accepts one of "EUR" or "NAM"\n
    Output is list of dicts, each structured as {'Airline': x, 'Code': y, 'Type': z}
    MUST HAVE ALREADY SHARED SHEET WITH OWNER OF CREDENTIAL
    """
    from google.oauth2.service_account import Credentials
    import gspread
    from pandas import DataFrame as pdDataFrame
    # ed-cm-caranalytics-dev@appspot.gserviceaccount.com
    def __init__(self, spreadhseet, credentials='credentials.json'):
        self.scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        self.json_credentials = self.Credentials.from_service_account_file(credentials, scopes=self.scope)
        self.gc = self.gspread.authorize(self.json_credentials)
        self.gss = self.gc.open(spreadhseet)

    def getWS(self, WhichProcess, asDF=True):
        wks = self.gss.worksheet(WhichProcess)
        wksvalues=wks.get_all_values()
        if not asDF:
            return [[x, y] for x,y in wksvalues[1::]]
        else:
            return self.pdDataFrame(wksvalues[1:], columns=wksvalues[0])

    def Airlines(self, WhichProcess, asDF=True):
        wks = self.gss.worksheet(WhichProcess)
        wksvalues=wks.get_all_values()
        if asDF:
            CompetenciaList = self.pdDataFrame(wksvalues[1::], columns=wksvalues[0])
        else:
            CompetenciaList = [{'Airline': x, 'Code': y, 'Type': z} for x,y,z in wks.get_all_values()[1::]]

        # CompetenciaList = {y: {'Airline': x, 'Tipo': z} for x,y,z in wks.get_all_values()}
        return CompetenciaList

    def Mails(self, Testing):
        wks = self.gss.worksheet("Correos")
        return wks.get_all_values() if not Testing else ["equipo.planificacionrmcargo@gmail.com"]

    def Hauls(self):
        wks = self.gss.worksheet("Hauls")
        HaulsDict = {x: y for x,y in wks.get_all_values()}
        return HaulsDict

    def Geo(self):
        wks = self.gss.worksheet("Geographic")
        Geographic=wks.get_all_values()
        return self.pdDataFrame(Geographic[1:], columns=Geographic[0])

    def JoinWachas(self, guachas, AllAirlines):
        guachitas=guachas.merge(AllAirlines[["NAME", "CODE"]], how="left", on="CODE")[["NAME", "CODE", "TYPE", "ARCFT", "REGNUM"]]
        return guachitas.values.tolist()


# endregion
# ---------------------------------------------------------------------------------------

# region: Post to Slack -----------------------------------------------------------------

from slacker import Slacker
from slack_progress import SlackProgress # https://github.com/bcicen/slack-progress
class SlackMsg:
    def __init__(self, cnl="ruteo-diariofr24"):
        from keys import slack_token
        # QUESTION ... who's api token is this?
        self.api_token = slack_token
        self.channel = cnl
        self.username = "HAL 9000"
        self.slack = Slacker(self.api_token)

    def Post(self, text):
        self.slack.chat.post_message(channel=self.channel, text=text, username=self.username)
        print(text)

    def Hello(self, monthDay, hourMins):
        smsg = "{}\n\n{}: Corriendo proceso Ruteo diario a las {}.\n...".format("- "*20, monthDay, hourMins)
        self.Post(text=smsg)

    def SlackProgressBar(self, total):
        self.intotal = total
        self.sp = SlackProgress(self.api_token, self.channel)
        self.pbar = self.sp.new(total=self.intotal)

    def SlackProgressBarUpdater(self, posupdate):
        self.pbar.pos = posupdate

# endregion
# ---------------------------------------------------------------------------------------

# region: GetInfo from fr24 -------------------------------------------------------------


# define function that gets RegNum info for each airline

class DefaultHeaders:
    def __init__(self):
        UserAgents=[
            "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:52.0) Gecko/20100101 Firefox/52.0", 
            "Mozilla/5.0 (Windows; U; Windows NT 5.1; cs; rv:1.9.0.8) Gecko/2009032609 Firefox/3.0.8"
            ]
        self.UserAgent=UserAgents[processTime().intDay%2]
        self.defaultheaders = {"User-Agent": self.UserAgent}


import requests
from bs4 import BeautifulSoup
# Disable warnings
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from pandas import DataFrame as pdDataFrame

class fr24ws:
    """
    modules: GetRegNums(), LoopOverAlnCodes(), Flights2DataFrame(), Geo2DataFrame()
    """
    def __init__(self, Testing=False):
        self.defaultHeaders = DefaultHeaders().defaultheaders
        self.subprcstime = processTime()
        self.pageTimeStamp = self.subprcstime.NowTimeStamp
        self.fr24wsSlackMsg = SlackMsg(cnl="code-tests") if Testing else SlackMsg()

    def GetRegNums(self, airlineinfo):
        """
        airlineinfo must be a dict as such: {'Airline': 'ABX', 'Code': 'gb-abx', 'Type': 'cao'}
        'Code' key's value MUST be a fr24 established code for desired airline (generally formatted as iata-icao: xx-xxx)
        """
        headers=self.defaultHeaders
        arlrgnloopstart=dt.datetime.now()
        airline=airlineinfo['Airline']
        aircode=airlineinfo['Code']
        airtype=airlineinfo['Type']
        AirlineRegNumList = []
        url0 = 'https://www.flightradar24.com/data/airlines/'+aircode+'/fleet'
        req1 = requests.get(url0, headers=headers, verify=False)
        soup = BeautifulSoup(req1.content, 'html.parser')
        dltag = soup.find('dl', attrs={'id': 'list-aircraft'})
        try:
            AircraftModelList = dltag.find_all('dt', attrs={'class': None})
            TableBodyList = dltag.find_all('tbody')
            lenAircraftModelList=len(AircraftModelList)
            for i in range(lenAircraftModelList):
                aircraftmodel = AircraftModelList[i].find('div').text
                tbodyList = [x.text.strip() for x in TableBodyList[i].find_all('a', attrs={'class': 'regLinks'})]
                for regnum in tbodyList:
                    AirlineRegNumList.append([airline, aircode, airtype, aircraftmodel, regnum])
        except:
            pass
        smsg = f"Finished getting {airline} ({aircode}) in {self.subprcstime.SubProcRunTime(arlrgnloopstart)}"
        print(smsg)
        return AirlineRegNumList

    # Get RegNums for each Airline by code
    def LoopOverAlnCodes(self, Airlinefr24CodeDict):
        """
        Get RegNums for each Airline by code (generally iata-icao)
        """
        smsg = "Getting Registration Numbers for required Airlines..."
        self.fr24wsSlackMsg.Post(text=smsg)
        AllRegNumsList = []
        for AirlineDict in Airlinefr24CodeDict:
            AllRegNumsList.extend(self.GetRegNums(AirlineDict))
        # smsg = f"Finished scrapping Aircraft Rgn in {processTime(compListloopstart)} seconds"
        smsg = f"Finished scrapping Aircraft Rgn"
        self.fr24wsSlackMsg.Post(text=smsg)
        return AllRegNumsList

    # Get flight info per RegNum
    def GetFlightInfo(self, AircraftRegNums):
        """
        Get flight info per RegNum(s), AircraftRegNums arg must be a list.
        returns datatable, geoListOfLists
        """
        from keys import fr_token
        headers=self.defaultHeaders
        # headers=defaultHeaders
        spageTimeStamp=self.pageTimeStamp
        # spageTimeStamp=pageTimeStamp

        LoopTimeStart=dt.datetime.now()
        datatable = []
        geoListOfLists = []
        ptoken = fr_token
        loopLen = len(AircraftRegNums)
        # innerLoopSlackMsg=SlackMsg()
        smsg="Looping over Registration Numbers..."
        self.fr24wsSlackMsg.Post(text=smsg)
        self.fr24wsSlackMsg.SlackProgressBar(total=loopLen)
        num=0
        for RegNumInfo in AircraftRegNums:
            iRegNumTimeStart=dt.datetime.now()
            num+=1
            # numTimer = datetime.now()
            # RegNumInfo=AircraftRegNums[0]
            aln, cde, typ, act, rgn = RegNumInfo
            url1 = 'https://www.flightradar24.com/data/aircraft/' + rgn
            url2 = f'https://api.flightradar24.com/common/v1/flight/list.json?query={rgn}&fetchBy=reg&page=1&pk=&limit=100&token={ptoken}&timestamp={str(spageTimeStamp)}'
            s = requests.session()
            r = s.get(url1, headers=headers, verify=False)
            cookie = r.cookies.get_dict()
            headersFull = {"User-Agent": headers["User-Agent"], "Content-Type": "application/json", "x-fetch": "true"}
            response = None
            while response is None:
                try:
                    response = s.get(url2, cookies=cookie, headers=headersFull, verify=False).json()
                except:
                    pass
            try:
                data = response['result']['response']['data']
            except KeyError:
                data = None
            if data != None:
                # row=data[0]
                for row in data:
                    # initialize variables on each loop (clean them)
                    # get data from json
                    callsn = row['identification']['callsign'] # callsign
                    fltnum = row['identification']['number']['default'] # flight number
                    statusRaw = row['status']['text'] # status
                    # clean and separate status data
                    status = reStatus(statusRaw)
                    statusTime = reTime(statusRaw)
                    # utc of departure
                    deptime1 = row['time']['real']['departure']
                    deplocaltime = ToTimeIfPossible(deptime1) # None if deptime1 == None else dt.datetime.fromtimestamp(deptime1).strftime('%Y-%m-%d %H:%M:%S') 
                    # utc of event
                    arrtime1 = row['time']['real']['arrival']
                    arrlocaltime = ToTimeIfPossible(arrtime1) # None if arrtime1 == None else dt.datetime.fromtimestamp(arrtime1).strftime('%Y-%m-%d %H:%M:%S')
                    # Origin info
                    try:
                        orginfo = row['airport']['origin']
                        orgato = orginfo['code']['iata']
                        orgctrycode = orginfo['position']['country']['code']
                        orgoffset = orginfo['timezone']['offset']
                        deptimeUTC = ToTimeIfPossible(deptime1 - orgoffset) # dt.datetime.fromtimestamp(deptime1 - orgoffset).strftime('%Y-%m-%d %H:%M:%S')
                    except TypeError:
                        orgato = None
                        orgctrycode = None
                        orgoffset = None
                        deptimeUTC = None
                    # Destino info
                    try:
                        desinfo = row['airport']['destination']
                        desato = desinfo['code']['iata']
                        desctrycode = desinfo['position']['country']['code']
                        desoffset = desinfo['timezone']['offset']
                        arrtimeUTC = ToTimeIfPossible(arrtime1 - desoffset) # None if arrtime1 == None else dt.datetime.fromtimestamp(arrtime1 - desoffset).strftime('%Y-%m-%d %H:%M:%S') # Accepts None for Estimated arrival cases
                    except TypeError:
                        desato = None
                        desctrycode = None
                        desoffset = None
                        arrtimeUTC = None
                    # list with info to append to "datatable" (list of lists, then change to a DF)
                    flightLoopData = [aln, typ, act, rgn, callsn, fltnum, orgato, desato, deptimeUTC, arrtimeUTC, deplocaltime, arrlocaltime, status, statusTime, ]
                    datatable.append(flightLoopData)
                    geoListOfLists.extend((orginfo, desinfo))
            # info to print to console, just to know where the process is at
            # rgnTime = processTime(numTimer)
            # totalTime = processTime(startTimer, 'f')
            print(f"{num}/{loopLen} in {self.subprcstime.SubProcRunTime(iRegNumTimeStart)} seconds, total looping time: {self.subprcstime.SubProcRunTime(LoopTimeStart)}")
            self.fr24wsSlackMsg.SlackProgressBarUpdater(posupdate=round(num/loopLen*100,2))
        smsg = f"Finished Looping in {self.subprcstime.SubProcRunTime(LoopTimeStart, format='miau')}"
        self.fr24wsSlackMsg.Post(text=smsg)
        return datatable, geoListOfLists

    # write Airline Flight info to a DataFrame 
    def Flights2DataFrame(self, datatable):
        """
        Flight Info: write as DataFrame
        """
        AirIteCols = [
            'Airline',
            'Type',
            'Aircraft',
            'RegNum',
            'Callsign',
            'FlightNum',
            'Org',
            'Des',
            'DepartureUTC',
            'ArrivalUTC',
            'DepartureLT',
            'ArrivalLT',
            'Status',
            'StatusTime',
            ]
        return pdDataFrame(datatable, columns=AirIteCols) 

    # write Geo info to a DataFrame
    def Geo2DataFrame(self, geoListOfLists):
        # loop over geoinfo and write as dataframe
        preGeoDF = []
        for row in geoListOfLists:
            if row:
                atoname = row.get('name')
                iata = row.get('code').get('iata')
                icao = row.get('code').get('icao')
                latitude = row.get('position').get('latitude')
                longitude = row.get('position').get('longitude')
                city = row.get('position').get('region').get('city')
                country = row.get('position').get('country').get('name')
                ctrycode = row.get('position').get('country').get('code')
                tzname = row.get('timezone').get('name')
                tzabbr = row.get('timezone').get('abbr')
                tzabbrname = row.get('timezone').get('abbrName')
                tzOffset = row.get('timezone').get('offset')
                preGeoDF.append([atoname, iata, icao, latitude, longitude, city, country, ctrycode, tzname, tzabbr, tzabbrname, tzOffset,])

        GeoCols = [
            'atoname',
            'iata',
            'icao',
            'latitude',
            'longitude',
            'city',
            'country',
            'ctrycode',
            'tzname',
            'tzabbr',
            'tzabbrname',
            'tzOffset',
            ]

        GeoFrame = pdDataFrame(preGeoDF, columns=GeoCols).drop_duplicates(subset="iata").reset_index(drop=True)
        return GeoFrame



# endregion
# ---------------------------------------------------------------------------------------

# region: Improve Scraped Info, load to GBQ and drop duplicates -------------------------

class Clean2GBQload:
    from google.cloud import bigquery
    from google.oauth2.service_account import Credentials

    _scope = ["https://www.googleapis.com/auth/cloud-platform"]
    _credentials = Credentials.from_service_account_file('credentials.json', scopes=_scope)
    _client = bigquery.Client(credentials=_credentials, project='ed-cm-caranalytics-dev')

    _miau={'South America': 'SAM', 'North America': 'NAM', 'Asia': 'ASIA', 'Oceania': 'OCE', 'Europe': 'EUR', 'Africa': 'AFR'}
    _contscols=['Two_Letter_Country_Code', 'Continent_Name']
    _contscolsrename=['ctrycode', 'contname']

    _AIcolumntypes = {
        'Airline': 'object',
        'Type': 'object',
        'Aircraft': 'object',
        'RegNum': 'object',
        'Callsign': 'object',
        'FlightNum': 'object',
        'Org': 'object',
        'Des': 'object',
        'DepartureUTC': 'datetime64[ns]',
        'ArrivalUTC': 'datetime64[ns]',
        'DepartureLT': 'datetime64[ns]',
        'ArrivalLT': 'datetime64[ns]',
        'Status': 'object',
        'StatusTime': 'object',
        'OrgCont': 'object',
        'DesCont': 'object',
        'OrgDesCont': 'object',
        'ScrapeDateId': 'datetime64[ns]',
    }

    def __init__(self, Geo, Conts, AircraftItinerario, Testing=False):
        contssub=Conts[self._contscols]
        contssub.columns=self._contscolsrename
        self.GeoTable=Geo.merge(contssub, how='left', on=self._contscolsrename[0])
        self.GeoTable['contabbr']=[self._miau[x] for x in self.GeoTable['contname']]
        self.contcross = self.GeoTable[['iata', 'contabbr']]

        AircraftItinerario2=AircraftItinerario\
            .merge(self.contcross.rename(columns={'iata': 'Org', 'contabbr': 'OrgCont'}), how='left', on='Org',)\
            .merge(self.contcross.rename(columns={'iata': 'Des', 'contabbr': 'DesCont'}), how='left', on='Des',)\
            .assign(
                OrgDesCont=lambda x: x.OrgCont+"-"+x.DesCont,
                ScrapeDateId=processTime().scrapedateid,
            )
        cleanfilt = (AircraftItinerario2.Org.notnull() & AircraftItinerario2.Des.notnull() & [x in ['Landed', 'Estimated'] for x in AircraftItinerario2.Status])

        self.FlightsInfo = AircraftItinerario2[cleanfilt].copy().astype(self._AIcolumntypes)
        
        self.c2bqslackMsg = fr24ws(Testing=Testing).fr24wsSlackMsg 

    def Geo2GBQ(self):
        # Upload to GBQ
        dataset_name = 'parametricas_PAC'
        table_name = 'PAC_Geografica'
        table_ref = self._client.dataset(dataset_name).table(table_name)

        job_config = self.bigquery.LoadJobConfig()
        job_config.create_disposition = 'CREATE_IF_NEEDED'
        job_config.write_disposition = self.bigquery.WriteDisposition.WRITE_APPEND
        job_config.autodetect = True
        self._client.load_table_from_dataframe(dataframe=self.GeoTable, destination=table_ref, job_config=job_config).result()

        smsg="Uploaded GeoTable to bq"
        self.c2bqslackMsg.Post(text=smsg)

        DropDups_job_config = self.bigquery.QueryJobConfig(destination=table_ref)
        DropDups_job_config.write_disposition = self.bigquery.WriteDisposition.WRITE_TRUNCATE

        sql = """
        WITH ranked_messages AS (
        SELECT m.*, ROW_NUMBER() OVER (PARTITION BY iata, ctrycode) AS rn
        FROM `parametricas_PAC.PAC_Geografica` AS m
        )
        SELECT * EXCEPT(rn) FROM ranked_messages WHERE rn = 1;
        """

        # Start the query, passing in the extra configuration.
        query_job = self._client.query(sql, job_config=DropDups_job_config)  # Make an API request.
        query_job.result()  # Wait for the job to complete.

        smsg="Dropped GeoTable duplicates"
        self.c2bqslackMsg.Post(text=smsg)

    def Flights2GBQ(self):
        smsg="Uploading Flight data to BQ"
        self.c2bqslackMsg.Post(text=smsg)

        dataset_name = 'fr24WSxRgn'
        table_name = 'df_fr24ws_VoladoCompetencia'

        # AircraftScrapeSchema
        schema = [
            # bigquery.SchemaField('Airline', 'STRING', mode='REQUIRED'),
            self.bigquery.SchemaField('Airline', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('Type', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('Aircraft', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('RegNum', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('Callsign', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('FlightNum', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('Org', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('Des', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('DepartureUTC', 'TIMESTAMP', mode='NULLABLE'),
            self.bigquery.SchemaField('ArrivalUTC', 'TIMESTAMP', mode='NULLABLE'),
            self.bigquery.SchemaField('DepartureLT', 'TIMESTAMP', mode='NULLABLE'),
            self.bigquery.SchemaField('ArrivalLT', 'TIMESTAMP', mode='NULLABLE'),
            self.bigquery.SchemaField('Status', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('StatusTime', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('OrgCont', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('DesCont', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('OrgDesCont', 'STRING', mode='NULLABLE'),
            self.bigquery.SchemaField('ScrapeDateId', 'TIMESTAMP', mode='NULLABLE'),
            ]


        dataset_ref = self._client.dataset(dataset_name)
        table_ref = dataset_ref.table(table_name)
        table = self.bigquery.Table(table_ref, schema=schema)
        ## create table if doesn't exist, delete
        self._client.create_table(table, exists_ok=True)
        # _client.delete_table(table_ref)

        job_config = self.bigquery.LoadJobConfig()
        job_config.create_disposition = 'CREATE_IF_NEEDED'
        job_config.write_disposition = self.bigquery.WriteDisposition.WRITE_APPEND
        job_config.schema = schema
        self._client.load_table_from_dataframe(dataframe=self.FlightsInfo, destination=table_ref, job_config=job_config).result()

        smsg="Upload succesfull, now dropping duplicates"
        self.c2bqslackMsg.Post(text=smsg)

        # Overwrite table with most recent unique results
        # https://cloud.google.com/bigquery/docs/writing-results

        # TODO(developer): Construct a BigQuery _client object.
        # _client = bigquery.Client()
        # TODO(developer): Set table_id to the ID of the destination table.
        # table_id = "your-project.your_dataset.your_table_name"
        table_id = self._client.dataset(dataset_name).table(table_name)
        job_config = self.bigquery.QueryJobConfig(destination=table_id)
        job_config.write_disposition = self.bigquery.WriteDisposition.WRITE_TRUNCATE

        sql = f"""
        WITH ranked_messages AS (
        SELECT m.*, ROW_NUMBER() OVER (PARTITION BY Airline, Aircraft, RegNum, Callsign, FlightNum, Org, Des, DepartureUTC ORDER BY ScrapeDateId DESC) AS rn
        FROM `{dataset_name}.{table_name}` AS m
        )
        SELECT * EXCEPT(rn) FROM ranked_messages WHERE rn = 1;
        """

        # Start the query, passing in the extra configuration.
        query_job = self._client.query(sql, job_config=job_config)  # Make an API request.
        query_job.result()  # Wait for the job to complete.

        smsg="Duplicates dropped, most recent data withheld"
        self.c2bqslackMsg.Post(text=smsg)


# endregion
# ---------------------------------------------------------------------------------------

# region: query and import BQ data ------------------------------------------------------ 

class queryGBQ2DF:
    from google.cloud import bigquery
    from google.oauth2.service_account import Credentials
    _scope = ["https://www.googleapis.com/auth/cloud-platform"]
    _credentials = Credentials.from_service_account_file('credentials.json', scopes=_scope)
    _client = bigquery.Client(credentials=_credentials, project='ed-cm-caranalytics-dev')

    def getQuery(self, query):
        """
        Runs query string to predetermined gcp bq project, returns results as dataframe.
        """
        return self._client.query(query).result().to_dataframe()

    def GetTestQuery(self):
        testquery="""
            SELECT * FROM `ed-cm-caranalytics-dev.fr24WSxRgn.df_fr24ws_VoladoCompetencia` 
            WHERE DepartureUTC > TIMESTAMP_ADD(CURRENT_TIMESTAMP(), INTERVAL -7*6 DAY)
            AND RegNum = 'XA-LFR'
            LIMIT 1000
        """
        return self._client.query(testquery).result().to_dataframe()

    def GetByRegNum(self, RegNum="N781AV", weeks=1):
        query=f"""
            SELECT * FROM `ed-cm-caranalytics-dev.fr24WSxRgn.df_fr24ws_VoladoCompetencia` 
            WHERE DepartureUTC > TIMESTAMP_ADD(CURRENT_TIMESTAMP(), INTERVAL -7*{weeks} DAY)
            AND RegNum = '{RegNum}'
            LIMIT 1000
        """
        return self._client.query(query).result().to_dataframe()



    def getRecentFlightData(self):
        """
        Runs predetermined query to predetermined gcp bq project, returns results as dataframe.
        Selects everythin from df_fr24ws_VoladoCompetencia where DepartureUTC is greater than today - 7*6 days.
        """
        query=f"""
                SELECT * FROM `ed-cm-caranalytics-dev.fr24WSxRgn.df_fr24ws_VoladoCompetencia` 
                WHERE DepartureUTC > TIMESTAMP_ADD(CURRENT_TIMESTAMP(), INTERVAL -7*6 DAY)
            """
        return self.getQuery(query)

# endregion
# ---------------------------------------------------------------------------------------

# region: Flight Router Functions -------------------------------------------------------

class Routing:

    deses = GetGoogleSheetInfo('fr24ws').Hauls()
    import regex as re2
    from pandas import DataFrame as pdDataFrame

    DictTime = DictTime()
    thisweek = processTime().thisWeek
    llvalues=['W-7', 'W-6', 'W-5', 'W-4', 'W-3', 'W-2', 'W-1', 'W']
    llkeys=list(range(thisweek-7, thisweek+1))

    def LHorSH(self, rtg):
        return 'LH' if ('LH' in [self.deses.get(x) for x in self.re2.findall(r'[A-Z]{3}', rtg, overlapped=True)]) else 'SH'

    def _RouteMeUp(self, FlightInfoDF, OrgDes):
        AvPax = True if OrgDes == "BOG-SAM" else False
        OrgCont, DesCont = OrgDes.split("-")
        ValidRegNums = self._FindRegNums(FlightInfoDF, OrgCont, DesCont, AvPax)

        ruteoMotherList = []
        for regnum in ValidRegNums:
            castor = self._ReturnRegNumInfo(FlightInfoDF, regnum)
            LenCastor = len(castor)
            ruteoMotherList += self._RegNumRouteByOrgDes(OrgCont, DesCont, castor, regnum)

        rmup=ruteoMotherList

        RuteoCols = ['OrgDesCont', 'Competidor', 'MatriculaAvion', 'Haul', 'WeekDep', 'DoW', 'HoraDep', 'Ruteo', 'FechaDep', 'FechaArr', 'RuteoCont'] 
        RuteoDF = self.pdDataFrame(rmup, columns=RuteoCols).sort_values(['Competidor', 'MatriculaAvion', 'FechaDep', ])
        # RuteoDF = pdDataFrame(ruteoMotherList, columns=RuteoCols).sort_values(['Competidor', 'MatriculaAvion', 'FechaDep', ])
        RuteoDF['RelWeek'] = self.DictTime.RelativeWeek(RuteoDF.FechaDep)
        # RuteoDF['RelWeek'] = DictTime.RelativeWeek(RuteoDF.FechaDep)
        return RuteoDF 

    def _FindRegNums(self, FlightInfoDF, OrgCont, DesCont, AvPax):
        """
        Gets list of RegNums that comply to given Org-Des (formatted as org-des)
        if AvPax assigned to True, there is no need to provide values for Org-Des
        """
        if AvPax:    
            atosCO=["APO","AUC","AXM","BSC","EJA","BAQ","BOG","GYM","BGA","BUN","CLO","CTG","CRC","CZU","CUC","EYP","FLA","GIR","GPI","IBE","LET","MZL","MQU","EOH","MDE","MVP","MTR","NVA","PSO","PEI","PPN","PVA","PUU","PCR","UIB","RCH","ADZ","SJE","SVI","SMR","RVE","TME","TLU","TCO","VUP","VVC",]
            cond1 = f"(Org == '{OrgCont}' and DesCont == '{DesCont}' and Des not in {atosCO} and Airline in ['AviancaCargo', 'Avianca'])"
        else:
            cond1 = f"(OrgCont == '{OrgCont}' and DesCont == '{DesCont}')"
            
        ValidRegNums = list(set(FlightInfoDF.query(f"{cond1}").RegNum))

        return ValidRegNums

    def _ReturnRegNumInfo(self, FlightInfoDF, regnum):
        """
        Returns df sorted by DepartureUTC, where RegNum matches the provided argument. 
        """
        # make loop temp table named castor to loop over each register number
        castor = FlightInfoDF.query(f" RegNum == '{regnum}' ").sort_values(by=['DepartureUTC']).reset_index(drop=True)
        return castor

    def _RegNumRouteByOrgDes(self, OrgCont, DesCont, castor, regnum):
        """
        Routes flights done by an airplane that match OrgDesCont input
        castor is subset of FlightInfoDF, one regnum sorted by DepartureUTC
        """
        
        # step 1: find flights that comply to OrgCont-DesCont
        LenCastor = len(castor) - 1

        aln = castor.iloc[0]['Airline']
        acft = castor.iloc[0]['Aircraft']
        rgninfo = [OrgCont+"-"+DesCont, aln, regnum]

        ruteoList = []

        # leftover testing
        # castor[[True if x == "NAM-SAM" else False for x in castor.OrgDesCont]] # 13, 24,28, 31
        # j=13

        for j in list(range(LenCastor)):
            jrow=castor.iloc[j]
            jNxtrow=castor.iloc[j+1]

            jOrgDesCont=jrow.OrgDesCont
            jNxtOrgDesCont=jNxtrow.OrgDesCont
            if (jOrgDesCont == OrgCont+"-"+DesCont):
                # use p to count distance to end of logic
                p=1
                while (j + p < LenCastor) and (jNxtOrgDesCont != DesCont+"-"+OrgCont) and (p < 5):
                    p +=1 
                rStart, rEnd = j, j+p # Org-Des up to exiting Des-Org (routing start, routing end)
                leading = (0 if rStart == 0 else rStart-1)
                trailing = (rEnd + 1 + 1)

                depy = jrow.DepartureLT
                dep = self.DictTime.Tstmp2Str(depy, 'ymdhms') # dep is departure from Org LT to Des
                hms = self.DictTime.Tstmp2Str(depy, "hms")
                dow = self.DictTime.DoW(depy) # day of week (LUN-DOM)
                dwk = self.DictTime.WeekNum(depy) # weeknum of dep lt
                arvy = jrow.ArrivalLT
                arv = self.DictTime.Tstmp2Str(arvy, 'ymdhms') # arv is arrival at Des LT

                castor2 = castor.iloc[leading:trailing].copy()
                # restructure elements for easy looping
                zippy = list(zip(castor2.Org, castor2.Des))
                zippyCont = list(zip(castor2.OrgCont, castor2.DesCont))
                LenZippy = len(zippy)

                # add leading airport (airport before org-des flight)
                rtg, rtgCont = zippy[0][0].lower(), zippyCont[0][0].lower()
                for i in range(LenZippy-1):
                    if zippy[i][1] == zippy[i+1][0]:
                        rtg = rtg+"-"+zippy[i+1][0]
                        rtgCont = rtgCont+"-"+zippyCont[i+1][0]
                    else:
                        # if error in arrival airport in i isn´t departure airport i i+1, assume missing data error and "join the pieces"
                        rtg = rtg+"-"+zippy[i][1]+"*-"+zippy[i+1][0]
                        rtgCont = rtgCont+"-"+zippyCont[i][1]+"*-"+zippyCont[i+1][0]
                # add trailing airport (airport after sam-nam flight)
                rtg = rtg+"-"+zippy[LenZippy-1][1].lower()
                # TODO LH definition 
                lhorsh = self.LHorSH(rtg)
                rtgCont = rtgCont+"-"+zippyCont[LenZippy-1][1].lower()
                rgninfo2 = [lhorsh, dwk, dow, hms, rtg, dep, arv, rtgCont] #
                ruteoList.append(rgninfo + rgninfo2)

        return ruteoList

    def _PivRoutes(self, RoutedTable):
        from pandas import Categorical as pdCategorical
        weeks_to_cols=dict(zip(self.llkeys, self.llvalues))
        rdf=RoutedTable[['OrgDesCont', 'Competidor', 'MatriculaAvion', 'Haul', 'RelWeek', 'DoW', 'HoraDep', 'Ruteo',]].copy().query(f"RelWeek in {self.llkeys}")
        rdf["Weeks2Cols"]=[weeks_to_cols.get(x, None) for x in rdf['RelWeek']]
        pivdreport=rdf\
            .pivot_table(
                index=['OrgDesCont', 'Competidor', 'MatriculaAvion', 'Haul', 'DoW', 'HoraDep'], 
                columns = ['Weeks2Cols'], 
                values='Ruteo',
                aggfunc=lambda x: ', '.join(x.unique()))\
            .reset_index()\
            .drop(columns=['HoraDep'])\
            .fillna("-")
        pivdreport['DoW'] = pdCategorical(pivdreport['DoW'], categories=["LUN","MAR","MIE","JUE","VIE","SAB","DOM",], ordered=True)
        pivdreport=pivdreport.sort_values(['OrgDesCont', 'Competidor', 'DoW', 'Haul',])
        return pivdreport

    def _SendTableFormat(self, RoutedAndPivd):
        from pandas import Series as pdSeries 
        grbygrp=['Competidor', 'Haul']
        prepy=pdDataFrame()
        CompHauls=tuple(RoutedAndPivd.groupby(grbygrp).size().reset_index()[grbygrp].itertuples(index=False, name=None))
        for Competidor, Haul in CompHauls:
            CompHaulLoop = RoutedAndPivd.loc[(RoutedAndPivd.Competidor == Competidor) & (RoutedAndPivd.Haul == Haul)]
            cols_to_weeks=dict(zip(self.llvalues[:2:-1], self.llkeys[:2:-1]))
            fakedict=dict()
            OrgDesCont=CompHaulLoop['OrgDesCont'].iloc[0]
            fakedict['OrgDesCont']=OrgDesCont
            fakedict['Competidor']=Competidor
            fakedict['Haul']=Haul
            for col, wk in cols_to_weeks.items():
                try:
                    chlcols=["DoW", col, "MatriculaAvion"]
                    b=CompHaulLoop.loc[CompHaulLoop[col] != "-"][chlcols]
                    a=[f"{x}: {y} ({z})" for x,y,z in zip(b[chlcols[0]], b[chlcols[1]], b[chlcols[2]])]
                    if a!= []:
                        fakedict[wk]=a
                except KeyError:
                    next

            prep=pdDataFrame(dict([ (k, pdSeries(v)) for k,v in fakedict.items() ]))
            prep['OrgDesCont']=OrgDesCont
            prep['Competidor']=Competidor
            prep['Haul']=Haul
            prepy = prepy.append(prep)

        srtdWeek=sorted(list((set(list(prepy))-set(['Competidor', 'Haul', 'OrgDesCont', ]))))[::-1]
        prepy=prepy.reset_index(drop=True)[['Competidor', 'Haul', 'OrgDesCont',]+srtdWeek]
        prepy.to_csv("processfiles/table.csv", index=False)
        return prepy

    def _RouteAndPiv(self, FlightInfoDF, OrgDesList):
        """
        Requiered input: FlightInfoDF in corresponding format, List of ORG-DES pairs as strings.
        """
        # Route inputs acordingly
        AppendedRMU=pdDataFrame()
        for OrgDes in OrgDesList:
            print(f"Routing {OrgDes}")
            AppendedRMU = AppendedRMU.append(self._RouteMeUp(FlightInfoDF, OrgDes)) # capture appending to df
        # piv AppendedRMU
        print("Pivoting Routed info")
        routedandpivd = self._PivRoutes(AppendedRMU)
        return routedandpivd

    def Run(self, FlightInfoDF, OrgDesList):
        RoutedAndPivd=self._RouteAndPiv(FlightInfoDF, OrgDesList)
        sendtable=self._SendTableFormat(RoutedAndPivd)
        return sendtable

# endregion
# ---------------------------------------------------------------------------------------

# format png for email ------------------------------------------------------------------
# region

def formatPNG(tableFrame, folder="processfiles/"):
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle

    # func for cleaning all data
    def DataCleaner(ws):
        for row  in ws.iter_rows(min_row=3):
            for cell in row:
                cell.value = None
                cell.border = None
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(color='000000', bold=False)

    def as_texto(value):
        if value is None:
            return ""
        return str(value)

    Comps_y_postas=[
        {'Airline': 'Atlas', 'desc': {'color': '1f497d', 'postas': ['SCL', 'VCP', 'MAO']}},
        {'Airline': 'SkyLeaseCargo', 'desc': {'color': 'ffbf00', 'postas': ['SCL', 'AGT']}},
        {'Airline': 'KalittaAir', 'desc': {'color': '000000', 'postas': ['SCL', 'AGT']}},
        {'Airline': 'AviancaCargo', 'desc': {'color': 'ff0000', 'postas': ['SCL', 'VCP', 'MAO']}},
        {'Airline': 'Korean', 'desc': {'color': '4f81bd', 'postas': ['SCL', 'VCP']}},
        {'Airline': 'TurkishAirlines', 'desc': {'color': 'ff4b4b', 'postas': []}},
    ]

    xlsxname=folder+"FormatoReporte.xlsx"
    pngname=folder+"ResumenReporte.png"

    if processTime().startTimer.weekday() == 0:
        column_selection = [0,1,2,4,5]
    else:
        column_selection = [0,1,2,3,4]

    wb = load_workbook(filename=xlsxname)
    ws = wb['Todos']
    DataCleaner(ws)
    thickblackborder = Side(border_style="thick", color = "000000") # styling for cells
    rcellstart = 0 # rcellstart will store from which rownumber to start building each tabe
    ws['B1'].value = processTime().todayfrmtd

    # cyp = Comps_y_postas[3] # use for testing followig loop
    for cyp in Comps_y_postas:

        Airline=cyp['Airline']
        CellColor=cyp['desc']['color'].upper()
        Postas=cyp['desc']['postas']

        dftest=tableFrame.iloc[:,column_selection]\
            .query(f"Competidor == '{Airline}' and Haul == 'LH' and OrgDesCont == 'NAM-SAM' ")\
            .reset_index(drop=True)
        
        if dftest.iloc[:, 3:5].isnull().all().all():
            continue
            
        rcellstart += 2 # give two rows headspace between tables

        WeeksSubset=dftest.iloc[:,-2:].copy()
        WeeksSubset=WeeksSubset.dropna(how='all').fillna("")

        dftest=dftest.iloc[WeeksSubset.index].fillna("")

        if bool(Postas):
            toques=[[ f"{Posty}: {sum(x.count(Posty) for x in WeeksSubset[colm])}" for Posty in Postas] for colm in list(WeeksSubset)]
        else: toques=""

        rowcount, colcount= dftest.shape
        header_values = list(dftest)

        rangerowcount=range(rowcount)
        rangecolcount=range(colcount)

        # Edit excel table Header
        for c in rangecolcount:
            # insert header name to header row
            # TODO maybe start from blank sheet and apply formatting as well?
            wscell=ws.cell(row=rcellstart, column=c+2)
            wscell.value = header_values[c]
            wscell.fill = PatternFill(fgColor=CellColor, fill_type='solid')
            wscell.font = Font(color='D9D9D9', bold=True)
            wscell.alignment = Alignment(horizontal='center', vertical='center')

            if c == 0:
                wscell.border = Border(left=thickblackborder, bottom=thickblackborder, top=thickblackborder)
            elif c == colcount - 1:
                wscell.border = Border(right=thickblackborder, bottom=thickblackborder, top=thickblackborder)
            else:
                wscell.border = Border(bottom=thickblackborder, top=thickblackborder)

        # edit table data
        for r in rangerowcount:
            for c in rangecolcount:
                # insert info by row, if first/last col apply or last row apply formats
                wsc=ws.cell(row=r+rcellstart+1, column=c+2)
                wsc.value = dftest.iloc[r, c]
                wsc.alignment = Alignment(horizontal='center', vertical='center')

                if c == 0:
                    if r == rowcount - 1:
                        wsc.border = Border(left=thickblackborder, bottom=thickblackborder)
                    else:
                        wsc.border = Border(left=thickblackborder)
                if c == colcount - 1:
                    if r == rowcount - 1:
                        wsc.border = Border(right=thickblackborder, bottom=thickblackborder)
                    else:
                        wsc.border = Border(right=thickblackborder)
                if r == rowcount - 1 and c not in [0, colcount - 1]:
                    wsc.border = Border(bottom=thickblackborder)

        # edit toque info
        for c in range(len(toques)):
            for r in range(len(Postas)):
                wsc=ws.cell(row=r+rcellstart+1+rowcount, column=c+5)
                wsc.value = toques[c][r]
                wsc.alignment = Alignment(horizontal='center', vertical='center')

        rcellstart += rowcount + 1 + len(Postas)

        # for col in ws.column_dimensions:
        #     col.column_dimensions.width = 20

    wb.save(xlsxname)
    wb.close()

    import excel2img
    rcellstart = rcellstart if rcellstart > 0 else 5
    excel2img.export_img(xlsxname, pngname, "Todos", f"A1:G{rcellstart}")

    fr24ws().fr24wsSlackMsg.Post(text="image written")


# endregion
# ---------------------------------------------------------------------------------------

# region: SendMails ---------------------------------------------------------------------

class SendEmails():
    import smtplib 
    import mimetypes
    from email import encoders
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.image import MIMEImage

    def __init__(self, ReceiverEmails, SenderEmail, SenderPwd):
        """
        ReceiverEmails must be python list datatype
        """
        folder="processfiles/"
        xlsxname=folder+"FormatoReporte.xlsx"
        pngname=folder+"ResumenReporte.png"
        fileToSend=folder+"table.csv"
        todayfrmtd=processTime().todayfrmtd

        # Authentication 
        s = self.smtplib.SMTP('smtp.gmail.com', 587) 
        s.starttls() 
        s.login(SenderEmail, SenderPwd) 
        # creates SMTP session 
        # start TLS for security 
        img_data = open(pngname, 'rb').read()

        message = self.MIMEMultipart("alternative")
        message["Subject"] = f"Ruteo Diario Competencia, {todayfrmtd}"
        message["From"] = SenderEmail
        message["To"] = ", ".join(ReceiverEmails)
        # convert both parts to MIMEText objects and add them to the MIMEMultipart message
        # part1 = MIMEText(tbl, "plain")
        emailtext=f"""Estimados,

        Adjuntamos archivo csv del ruteo NAM-SAM, Asia-NAM, EUR-SAM, EUR-NAM, OCE-SAM y BOG-SAM (Avianca) del día {todayfrmtd}. 
        Se incluye además png con resumen para NAM-SAM.
        Pueden visualizar el csv en excel haciendo click derecho -> abrir con Excel.

        Tener ojo con el conteo de los toques cuando hay asteriscos para la posta en el ruteo, estos son debido a discontinuidades en el dato desde flightradar.

        Saludos
        Equipo PAC
 
        """
        # print(emailtext)


        part1 = self.MIMEText(emailtext, "plain")
        message.attach(part1)
        attchImg = self.MIMEImage(img_data, name=pngname)
        message.attach(attchImg)

        ctype, encoding = self.mimetypes.guess_type(fileToSend)
        if ctype is None or encoding is not None:
            ctype = "application/octet-stream"

        maintype, subtype = ctype.split("/", 1)
        fp = open(fileToSend, "rb")
        attachment = self.MIMEBase(maintype, subtype)
        attachment.set_payload(fp.read())
        fp.close()
        self.encoders.encode_base64(attachment)
        attachment.add_header("Content-Disposition", "attachment", filename=fileToSend)
        message.attach(attachment)
        # sending the mail 
        s.sendmail(SenderEmail, ReceiverEmails, message.as_string())
        # terminating the session 
        s.quit() 

        smsg="Routed data sent by email\nProcess closed"
        fr24ws().fr24wsSlackMsg.Post(text=smsg)

# endregion
# ---------------------------------------------------------------------------------------

